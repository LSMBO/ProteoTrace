import openpyxl
from Bio import SeqIO
from ProteinGroup import ProteinGroup
from Protein import Protein
from Peptide import Peptide

class Proline:
    def __init__(self, excel_file_path, fasta_file, protein_description_list=[]):
        self.excel_file_path = excel_file_path
        self.fasta_file = fasta_file
        if protein_description_list:
            self.settings_sheet = []
            self.settings_sheet_col_name = []
            self.psms_from_protein_sets = []
            self.psms_from_protein_sets_col_name = []
            self.load_excel_file()
            self.sequences = self.read_fasta(fasta_file)
            protein_id_list = self.get_protein_id_list(protein_description_list)
            self.peptides = self.search_peptides(protein_id_list)
            self.proteins, self.protein_groups = self.search_proteins()

    def get_protein_id_list(self, protein_description_list):
        id_list = []
        for desc in protein_description_list:
            id_list.append(desc.split(' ')[0])
        return id_list
    
    def load_excel_file(self):
        try:
            workbook = openpyxl.load_workbook(self.excel_file_path)
            if 'Search settings and infos' in workbook.sheetnames:
                settings_worksheet = workbook['Search settings and infos']
                settings_sheet = self.worksheet_to_table(settings_worksheet)
                self.settings_sheet_col_name = settings_sheet[0]
                self.settings_sheet = settings_sheet[1:]
            if 'All PSMs from protein sets' in workbook.sheetnames:
                psms_worksheet = workbook['All PSMs from protein sets']
                psms_from_protein_sets = self.worksheet_to_table(psms_worksheet)
                self.psms_from_protein_sets_col_name = psms_from_protein_sets[0]
                self.psms_from_protein_sets = psms_from_protein_sets[1:]
            workbook.close()
        except FileNotFoundError:
            print(f"Error: Excel file not found. {self.excel_file_path}")

    def worksheet_to_table(self, worksheet):
        table = []
        for row in worksheet.iter_rows(values_only=True):
            table.append(list(row))
        return table
    
    def read_fasta(self, fasta_file):
        sequences = []
        for record in SeqIO.parse(fasta_file, "fasta"):
            if not record.id.startswith("#") and not record.id.startswith('CON'):
                entry = {}
                entry["id"] = record.id
                entry["description"] = record.description
                entry["sequence"] = str(record.seq)
                sequences.append(entry)
        return sequences


    def get_protein_ids_from_row(self, row):
        sameset_col_name_index = self.psms_from_protein_sets_col_name.index("samesets_accessions")
        subset_col_name_index = self.psms_from_protein_sets_col_name.index("subsets_accessions")
        sameset_list = row[sameset_col_name_index].split(';')
        subset_list = []
        protein_list = []
        if row[subset_col_name_index]:
            subset_list = row[subset_col_name_index].split(';')
        protein_list = [elem.lstrip(' >') for elem in sameset_list + subset_list]
        return protein_list
        
    def search_peptides(self, protein_list):
        peptides = []
        peptide_id_step = 0
        peptide_sequences = []
                    
        for row in self.psms_from_protein_sets:
            protein_ids = self.get_protein_ids_from_row(row)
            if any(protein_id in protein_list for protein_id in protein_ids):
                peptide_id = row[self.psms_from_protein_sets_col_name.index('peptide_id')]
                peptide_sequence = row[self.psms_from_protein_sets_col_name.index('sequence')]
                peptide_sequences.append(peptide_sequence)
                peptide_id = peptide_id - peptide_id_step
                peptides.append(Peptide(self.sequences, peptide_id, self.psms_from_protein_sets_col_name, row))
        return peptides
    
    
    def search_proteins(self):
        protein_rows = {}
        protein_group_rows = {}
                
        if ("samesets_accessions" in self.psms_from_protein_sets_col_name and "subsets_accessions" in self.psms_from_protein_sets_col_name and "protein_set_id" in self.psms_from_protein_sets_col_name):
            protein_set_id_index = self.psms_from_protein_sets_col_name.index("protein_set_id")
            sameset_col_name_index = self.psms_from_protein_sets_col_name.index("samesets_accessions")
            subset_col_name_index = self.psms_from_protein_sets_col_name.index("subsets_accessions")
            for row in self.psms_from_protein_sets:
                sameset_list = row[sameset_col_name_index].split(';')
                subset_list = []
                if row[subset_col_name_index]:
                    subset_list = row[subset_col_name_index].split(';')
                protein_list = [elem.lstrip(' >') for elem in sameset_list + subset_list]
                
                if row[protein_set_id_index] in protein_group_rows:
                    protein_group_rows[row[protein_set_id_index]].append(row)
                else:
                    protein_group_rows[row[protein_set_id_index]] = [row]
    
                for protein_id in protein_list:
                    if protein_id in protein_rows:
                        protein_rows[protein_id].append(row)
                    else:
                        protein_rows[protein_id] = [row]

        proteins = []
        protein_groups = []
        for protein_id, protein_rows_list in protein_rows.items():
            proteins.append(Protein(self.sequences, protein_id, self.psms_from_protein_sets_col_name, protein_rows_list, self.peptides))
        
        for protein_group_id, protein_group_rows_list in protein_group_rows.items():
            protein_groups.append(ProteinGroup(self.sequences, protein_group_id, self.psms_from_protein_sets_col_name, protein_group_rows_list, self.peptides))
        
        return proteins, protein_groups
    
    
    def get_settings_sheet(self):
        return self.settings_sheet
    
    def get_psms_from_protein_sets(self):
        return self.psms_from_protein_sets

    def get_total_number_of_peptides(self):
        return len(self.psms_from_protein_sets) - 1 


